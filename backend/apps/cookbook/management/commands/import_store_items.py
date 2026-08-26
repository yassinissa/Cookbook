"""
Import per-SKU cost + conversion + packaging data from the source workbook's
"Store Items" sheet into Cookbook's ItemConversion / ItemConversionLine tables.

    python manage.py import_store_items --file "00 Store Items.xlsx"
    python manage.py import_store_items --file "..." --dry-run

Why Cookbook owns this: the deployed inventory-platform has a price for only a
handful of items and computes cost from purchase orders, while this sheet has a
real cost for ~2,111 SKUs plus the density conversions recipe costing needs.
Idempotent — re-running updates rows in place (matched on item_sku).
"""
from decimal import Decimal

from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone

from apps.cookbook.models import ItemConversion, ItemConversionLine, UnitScale
from apps.cookbook.models.item_supplement import CostSource
from apps.cookbook.parsing import parse_conversion_cell, to_decimal

# "Unit Scale" cell text -> UnitScale.code
BASE_UNIT_MAP = {'g': 'g', 'ml': 'ml', 'ea': 'EA', 'each': 'EA', 'pc': 'Pc', 'pcs': 'Pcs'}

# The 12 cooking-measure columns, by header text (matched case-insensitively,
# substring). Order matters only for reporting.
CONVERSION_HEADERS = [
    '1 tbs of ml or gram', '1 ts of ml or gram', '1/2 ts of ml or gram',
    '1/4 ts of ml or gram', '1/8 ts of ml or gram',
    '1 cup of ml or gram', '3/4 cup of ml or gram', '2/3 cup of ml or gram',
    '1/2 cup of ml or gram', '1/3 cup of ml or gram', '1/4 cup of ml or gram',
    '1/8 cup of ml or gram',
]
GRAM_EQUIV_HEADERS = [
    '1 tbs of ml into gram', '1 ts of ml into gram', '1/2 ts of ml into gram',
    '1/4 ts of ml into gram', '1/8 ts of ml into gram',
    '1 cup of ml into gram', '3/4 cup of ml into gram', '2/3 cup of ml into gram',
    '1/2 cup of ml into gram', '1/3 cup of ml into gram', '1/4 cup of ml into gram',
    '1/8 cup of ml into gram',
]


def _norm(s):
    return ' '.join(str(s or '').lower().split())


def _nonzero(d):
    """A 0 in these sheets means 'not filled in', not a real zero."""
    return d if (d is not None and d != 0) else None


class Command(BaseCommand):
    help = 'Import item cost + conversions from the "Store Items" sheet.'

    def add_arguments(self, parser):
        parser.add_argument('--file', required=True)
        parser.add_argument('--sheet', default='Store Items')
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **opts):
        import openpyxl

        wb = openpyxl.load_workbook(opts['file'], read_only=True, data_only=True)
        ws = wb[opts['sheet']]
        rows = ws.iter_rows(values_only=True)

        # ── locate the header row (first row containing a "Code" cell) ─────────
        header = None
        for r in rows:
            if any(_norm(c) == 'code' for c in r):
                header = [_norm(c) for c in r]
                break
        if header is None:
            self.stderr.write('No header row with a "Code" column found.')
            return

        def col(*names):
            for name in names:
                for i, h in enumerate(header):
                    if h == _norm(name) or _norm(name) in h:
                        return i
            return None

        idx = {
            'code': col('code'),
            'base_unit': col('unit scale'),
            'cost_per_scale': col('cost (kd)/ unit scale', 'cost (kd)/unit scale'),
            'order_unit': col('order unit', 'order \nunit'),
            'order_cost': col('cost (kd)/ order unit', 'cost (kd)/order unit'),
            'pack_qty': col('total unit', 'total \nunit'),
            'note': col('note to add'),
            'g_per_tbs': col('grams in\n1 tbs', 'grams in 1 tbs'),
            'g_per_piece': col('grams in\n1 piece', 'grams in 1 piece'),
            'pcs_per_pkt': col('pieces in\n1 pkt', 'pieces in 1 pkt'),
            'pcs_per_kg': col('pieces in\n1 kg', 'pieces in 1 kg'),
            'pcs_per_box': col('pieces or pkt\nin box', 'pieces or pkt in box'),
        }
        conv_cols = [col(h) for h in CONVERSION_HEADERS]
        geq_cols = [col(h) for h in GRAM_EQUIV_HEADERS]

        if idx['code'] is None:
            self.stderr.write('Could not find the Code column.')
            return

        units = {u.code: u for u in UnitScale.objects.all()}
        unit_by_lower = {u.code.lower(): u for u in UnitScale.objects.all()}

        stat = dict(rows=0, priced=0, with_conv=0, no_base_unit=0, skipped=0, lines=0)
        now = timezone.now()

        with transaction.atomic():
            for r in ws.iter_rows(min_row=1, values_only=True):
                code = r[idx['code']] if idx['code'] < len(r) else None
                if not code or _norm(code) in ('code', ''):
                    continue
                code = str(code).strip()
                if not code.startswith('B'):
                    continue
                stat['rows'] += 1

                base_txt = _norm(r[idx['base_unit']]) if idx['base_unit'] is not None else ''
                base_code = BASE_UNIT_MAP.get(base_txt)
                base_unit = units.get(base_code) if base_code else None

                cost = _nonzero(to_decimal(r[idx['cost_per_scale']])) if idx['cost_per_scale'] is not None else None
                order_cost = _nonzero(to_decimal(r[idx['order_cost']])) if idx['order_cost'] is not None else None
                pack_qty = _nonzero(to_decimal(r[idx['pack_qty']])) if idx['pack_qty'] is not None else None
                if cost is None and order_cost and pack_qty:
                    cost = (order_cost / pack_qty).quantize(Decimal('0.000001'))

                if base_unit is None and cost is not None:
                    stat['no_base_unit'] += 1

                defaults = {
                    'order_unit': str(r[idx['order_unit']] or '')[:20] if idx['order_unit'] is not None else '',
                    'order_cost': order_cost,
                    'pack_qty': pack_qty,
                    'note_to_add': str(r[idx['note']] or '') if idx['note'] is not None else '',
                    'grams_per_piece': _nonzero(to_decimal(r[idx['g_per_piece']])) if idx['g_per_piece'] is not None else None,
                    'pieces_per_pack': _nonzero(to_decimal(r[idx['pcs_per_pkt']])) if idx['pcs_per_pkt'] is not None else None,
                    'pieces_per_kg': _nonzero(to_decimal(r[idx['pcs_per_kg']])) if idx['pcs_per_kg'] is not None else None,
                    'pieces_or_pack_per_box': _nonzero(to_decimal(r[idx['pcs_per_box']])) if idx['pcs_per_box'] is not None else None,
                }
                if cost is not None and base_unit is not None:
                    defaults.update(base_unit=base_unit, cost_per_base_unit=cost,
                                    cost_source=CostSource.EXCEL_IMPORT, cost_updated_at=now)
                    stat['priced'] += 1

                # ── conversion lines ─────────────────────────────────────────
                parsed_lines = []
                for pos, ci in enumerate(conv_cols):
                    if ci is None or ci >= len(r):
                        continue
                    p = parse_conversion_cell(r[ci])
                    if not p:
                        continue
                    unit = unit_by_lower.get(p['unit'].lower())
                    if unit is None:
                        continue
                    geq = None
                    gi = geq_cols[pos] if pos < len(geq_cols) else None
                    if gi is not None and gi < len(r):
                        gp = parse_conversion_cell(r[gi])
                        geq = gp['quantity'] if gp else None
                    parsed_lines.append(dict(label=p['label'], quantity=p['quantity'],
                                             unit=unit, gram_equivalent=geq))
                if parsed_lines:
                    stat['with_conv'] += 1

                if opts['dry_run']:
                    continue

                obj, _ = ItemConversion.objects.update_or_create(item_sku=code, defaults=defaults)
                obj.lines.all().delete()
                for pl in parsed_lines:
                    ItemConversionLine.objects.create(item_conversion=obj, **pl)
                    stat['lines'] += 1

            if opts['dry_run']:
                transaction.set_rollback(True)

        wb.close()
        mode = '[DRY RUN — nothing saved]' if opts['dry_run'] else '[SAVED]'
        self.stdout.write(f'\n{mode}')
        for k, v in stat.items():
            self.stdout.write(f'  {k:14}: {v}')
