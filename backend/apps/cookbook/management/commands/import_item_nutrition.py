"""
Import per-SKU nutrition from the source workbook's "Nutritional Information"
sheet into Cookbook's ItemNutrition table.

    python manage.py import_item_nutrition --file "00 Store Items.xlsx"
    python manage.py import_item_nutrition --file "..." --dry-run

Values are per one unit_scale (per g / per ml / per each). Idempotent on item_sku.
"""
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.cookbook.models import ItemNutrition, UnitScale
from apps.cookbook.parsing import to_decimal

BASE_UNIT_MAP = {'g': 'g', 'ml': 'ml', 'ea': 'EA', 'each': 'EA', 'pc': 'Pc', 'pcs': 'Pcs'}

# ItemNutrition field  ->  header substring in the sheet
FIELD_HEADERS = {
    'calories':        'calories',
    'fat_g':           'fat (g)/unit',
    'protein_g':       'protein (g)/unit',
    'saturated_fat_g': 'saturated fat',
    'trans_fat_g':     'trans fat',
    'cholesterol_mg':  'cholesterol',
    'sodium_mg':       'sodium',
    'carbs_g':         'carbs',
    'fibers_g':        'fibers',
    'sugars_g':        'sugars (g)/unit',
    'added_sugars_g':  'added sugars',
}


def _norm(s):
    return ' '.join(str(s or '').lower().split())


class Command(BaseCommand):
    help = 'Import per-SKU nutrition from the "Nutritional Information" sheet.'

    def add_arguments(self, parser):
        parser.add_argument('--file', required=True)
        parser.add_argument('--sheet', default='Nutritional Information')
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **opts):
        import openpyxl

        wb = openpyxl.load_workbook(opts['file'], read_only=True, data_only=True)
        if opts['sheet'] not in wb.sheetnames:
            self.stderr.write(f'No "{opts["sheet"]}" sheet in {opts["file"]}.')
            return
        ws = wb[opts['sheet']]

        header = None
        for r in ws.iter_rows(values_only=True):
            if any(_norm(c) == 'code' for c in r):
                header = [_norm(c) for c in r]
                break
        if header is None:
            self.stderr.write('No header row with a "Code" column.')
            return

        def col(name):
            for i, h in enumerate(header):
                if _norm(name) in h:
                    return i
            return None

        i_code = col('code')
        i_unit = col('unit scale')
        i_check = col('nutrition check')
        nutrient_cols = {f: col(h) for f, h in FIELD_HEADERS.items()}

        units = {u.code: u for u in UnitScale.objects.all()}
        stat = dict(rows=0, saved=0, no_unit=0, all_zero=0)

        with transaction.atomic():
            for r in ws.iter_rows(min_row=1, values_only=True):
                code = r[i_code] if i_code is not None and i_code < len(r) else None
                if not code or not str(code).strip().startswith('B'):
                    continue
                code = str(code).strip()
                stat['rows'] += 1

                unit_txt = _norm(r[i_unit]) if i_unit is not None and i_unit < len(r) else ''
                unit = units.get(BASE_UNIT_MAP.get(unit_txt, ''))
                if unit is None:
                    stat['no_unit'] += 1
                    continue

                vals = {}
                for field, ci in nutrient_cols.items():
                    d = to_decimal(r[ci]) if ci is not None and ci < len(r) else None
                    vals[field] = d if d is not None else 0
                if not any(vals.values()):
                    stat['all_zero'] += 1
                check = str(r[i_check] or '')[:255] if i_check is not None and i_check < len(r) else ''

                if opts['dry_run']:
                    continue
                ItemNutrition.objects.update_or_create(
                    item_sku=code,
                    defaults={'unit_scale': unit, 'verification_notes': check, **vals},
                )
                stat['saved'] += 1

            if opts['dry_run']:
                transaction.set_rollback(True)

        wb.close()
        mode = '[DRY RUN — nothing saved]' if opts['dry_run'] else '[SAVED]'
        self.stdout.write(f'\n{mode}')
        for k, v in stat.items():
            self.stdout.write(f'  {k:10}: {v}')
