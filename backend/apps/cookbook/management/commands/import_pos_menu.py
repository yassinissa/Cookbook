"""
Import the POS modifier catalogue from a "Menu POS Applications" workbook
(e.g. "040 Dine (Menu POS Applications).xlsm").

    python manage.py import_pos_menu "040 Dine (Menu POS Applications).xlsm"
    python manage.py import_pos_menu "..." --branch Dine --dry-run

Reads three sheets:
  POSLavu Forced Modifier / POSLavu Optional Modifier
      → ModifierGroup + ModifierOption rows, grouped by the "…Modifier List"
        column. Marker / header rows are dropped; names are split EN/AR on the
        last "/" and de-padded.
  POSLavu Menu
      → for each item, matched to a DishRecipe by pos_item_name then name_en,
        a DishModifierGroup row is created for the named forced / optional
        group, plus a MenuLineModifier on the branch's Menu line if it exists.

Each option's `kind` (choice / type / addon / instruction) is *guessed* from
the sheet it came from + its price; every guess needs a human to confirm, and
no `pos_mods_string` is set — that comes from a real Lavu sales report, not
here. Idempotent: groups/options matched by name; re-run updates in place.
"""
import re
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import models, transaction

from apps.cookbook.models import (
    Branch, DishRecipe, Menu,
    ModifierGroup, ModifierOption, ModifierOptionKind, ModifierRole,
    ModifierSelection, DishModifierGroup, MenuLineModifier,
)

FORCED_SHEET = 'POSLavu Forced Modifier'
OPTIONAL_SHEET = 'POSLavu Optional Modifier'
MENU_SHEET = 'POSLavu Menu'

_TATWEEL = 'ـ'
_INSTRUCTION_PREFIXES = ('without ', 'no ', 'من غير ', 'بدون ')
_MARKER_RE = re.compile(
    r'(\*{3,}|selections?\s+only|egg\s+selection|^\d+\s+selection|- ?\d+ ?items?\b)',
    re.IGNORECASE,
)
# a forced-modifier option priced at or above this looks like a whole variant
# (its price replaces the base) rather than a small add-on
_VARIANT_PRICE = Decimal('2.000')


def _ascii(s):
    """Console-safe rendering for Windows cp1252 stdout."""
    return str(s).encode('ascii', 'replace').decode('ascii')


def _dec(v):
    try:
        return Decimal(str(v)) if v not in (None, '') else Decimal('0')
    except (InvalidOperation, TypeError):
        return Decimal('0')


def _clean(text):
    s = re.sub(r'\*+', '', str(text or '').replace(_TATWEEL, ''))
    return ' '.join(s.split()).strip()


def _split_name(raw):
    """'Chicken/دجاج' → ('Chicken', 'دجاج'); splits on the LAST '/'."""
    s = _clean(raw)
    if '/' in s:
        en, ar = s.rsplit('/', 1)
        return _clean(en), _clean(ar)
    return s, ''


def _is_marker(name_en, group_name):
    if not name_en:
        return True
    if name_en.strip().lower() == group_name.strip().lower():
        return True
    return bool(_MARKER_RE.search(name_en))


def _guess_kind(name_en, price, *, forced):
    low = name_en.lower()
    if any(low.startswith(p) for p in _INSTRUCTION_PREFIXES):
        return ModifierOptionKind.INSTRUCTION
    if forced:
        if price <= 0:
            return ModifierOptionKind.CHOICE
        return ModifierOptionKind.TYPE if price >= _VARIANT_PRICE else ModifierOptionKind.ADDON
    return ModifierOptionKind.ADDON if price > 0 else ModifierOptionKind.CHOICE


class Command(BaseCommand):
    help = 'Import POS modifier groups/options + dish attachments from a Menu POS Applications workbook.'

    def add_arguments(self, parser):
        parser.add_argument('file')
        parser.add_argument('--branch', default='Dine',
                            help='Branch name_en for the menu-line attachments (default: Dine).')
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **opts):
        import openpyxl

        try:
            wb = openpyxl.load_workbook(opts['file'], read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'File not found: {opts["file"]}')
        missing = [s for s in (FORCED_SHEET, OPTIONAL_SHEET, MENU_SHEET) if s not in wb.sheetnames]
        if missing:
            raise CommandError(f'Sheet(s) missing: {missing} — is this a Menu POS Applications workbook?')

        parsed = {}
        parsed.update(self._parse_modifier_sheet(wb[FORCED_SHEET], forced=True))
        # optional sheet: only add groups the forced sheet didn't define
        for name, data in self._parse_modifier_sheet(wb[OPTIONAL_SHEET], forced=False).items():
            parsed.setdefault(name, data)
        menu_rows = self._parse_menu_sheet(wb[MENU_SHEET])
        wb.close()

        self.r = {
            'groups': set(), 'options': 0, 'markers': 0,
            'dishes_matched': 0, 'dishes_unmatched': [], 'lines_linked': 0,
            'undefined_groups': set(),
        }
        self.r['markers'] = sum(g['markers'] for g in parsed.values())

        commit = not opts['dry_run']
        with transaction.atomic():
            self._load(parsed, menu_rows, opts['branch'], commit=commit)
            if not commit:
                transaction.set_rollback(True)
        self._print_report(dry=not commit)

    # ── parse ────────────────────────────────────────────────────────────
    def _parse_modifier_sheet(self, ws, *, forced):
        groups = {}
        for r in list(ws.iter_rows(values_only=True))[3:]:
            if not r or r[1] in (None, ''):
                continue
            group_name = _clean(r[3])
            if (not group_name or _MARKER_RE.search(group_name)
                    or 'this sheet is used' in group_name.lower()):
                continue
            bucket = groups.setdefault(
                group_name, {'options': [], 'markers': 0, 'forced': forced, 'max_select': None})
            name_en, name_ar = _split_name(r[1])
            if _is_marker(name_en, group_name):
                bucket['markers'] += 1
                m = re.match(r'(\d+)\s+selections?\s+only', name_en, re.IGNORECASE)
                if m:
                    bucket['max_select'] = int(m.group(1))
                continue
            price = _dec(r[2])
            bucket['options'].append({
                'name_en': name_en[:160], 'name_ar': name_ar[:160], 'price': price,
                'kind': _guess_kind(name_en, price, forced=forced),
            })
        return groups

    def _parse_menu_sheet(self, ws):
        out = []
        for r in list(ws.iter_rows(values_only=True))[3:]:
            if not r or r[1] in (None, ''):
                continue
            out.append({
                'name_en': _split_name(r[1])[0],
                'optional': _clean(r[4]) or None,
                'forced': _clean(r[5]) or None,
            })
        return out

    # ── load ─────────────────────────────────────────────────────────────
    def _load(self, parsed, menu_rows, branch_name, *, commit):
        group_objs = {}
        for gname, data in parsed.items():
            grp = self._upsert_group(gname, data, commit)
            group_objs[gname] = grp
            for i, opt in enumerate(data['options']):
                self._upsert_option(grp, opt, i, commit)

        menu = (Menu.objects.filter(branch__name_en__iexact=branch_name, is_active=True)
                .select_related('branch').first())
        branch = menu.branch if menu else Branch.objects.filter(name_en__iexact=branch_name).first()

        for row in menu_rows:
            attach = [(row['forced'], ModifierRole.FORCED), (row['optional'], ModifierRole.OPTIONAL)]
            if not any(g for g, _ in attach):
                continue
            dish = self._find_dish(row['name_en'], branch)
            if not dish:
                self.r['dishes_unmatched'].append(row['name_en'])
                continue
            self.r['dishes_matched'] += 1
            line = menu.lines.filter(dish=dish).first() if menu else None
            for gname, role in attach:
                if not gname:
                    continue
                grp = group_objs.get(gname)
                if grp is None:
                    grp = self._upsert_group(
                        gname, {'options': [], 'forced': role == ModifierRole.FORCED}, commit)
                    group_objs[gname] = grp
                    self.r['undefined_groups'].add(gname)
                self._attach(dish, line, grp, role, commit)

    def _upsert_group(self, name, data, commit):
        self.r['groups'].add(name)
        forced = data.get('forced', False)
        max_sel = data.get('max_select')
        if max_sel:
            selection, min_sel = ModifierSelection.MULTI, max_sel
        elif forced:
            selection, min_sel, max_sel = ModifierSelection.SINGLE, 1, 1
        else:
            selection, min_sel = ModifierSelection.MULTI, 0
        defaults = {'selection': selection, 'min_select': min_sel, 'max_select': max_sel}
        if not commit:
            return ModifierGroup(name_en=name[:120], **defaults)
        grp, _ = ModifierGroup.objects.get_or_create(name_en=name[:120], defaults=defaults)
        return grp

    def _upsert_option(self, grp, opt, sort, commit):
        self.r['options'] += 1
        if commit and grp.pk:
            ModifierOption.objects.update_or_create(
                group=grp, name_en=opt['name_en'],
                defaults={'name_ar': opt['name_ar'], 'price_delta': opt['price'],
                          'kind': opt['kind'], 'sort_order': sort},
            )

    def _find_dish(self, name_en, branch):
        qs = DishRecipe.objects.filter(is_current=True)
        if branch:
            qs = qs.filter(models.Q(branch_ref=branch) | models.Q(branch__iexact=branch.name_en))
        return (qs.filter(pos_item_name__iexact=name_en).first()
                or qs.filter(name_en__iexact=name_en).first())

    def _attach(self, dish, line, grp, role, commit):
        if not (commit and grp.pk):
            return
        DishModifierGroup.objects.get_or_create(dish=dish, group=grp, defaults={'default_role': role})
        if line is not None:
            _, created = MenuLineModifier.objects.get_or_create(
                menu_line=line, group=grp, defaults={'role': role})
            if created:
                self.r['lines_linked'] += 1

    # ── report ───────────────────────────────────────────────────────────
    def _print_report(self, *, dry):
        r, w = self.r, self.stdout.write
        w(self.style.MIGRATE_HEADING(
            '\n== POS menu import ' + ('(dry run - nothing saved) ==' if dry else '==')))
        w(f'  groups                {len(r["groups"])}')
        w(f'  options               {r["options"]}')
        w(f'  marker rows dropped    {r["markers"]}')
        w(f'  dishes matched         {r["dishes_matched"]}')
        w(f'  menu lines linked      {r["lines_linked"]}')
        if r['undefined_groups']:
            w(self.style.WARNING(
                f'  referenced by the menu but not defined: {sorted(r["undefined_groups"])}'))
        if r['dishes_unmatched']:
            w(self.style.WARNING(
                f'  dishes with a modifier but no matching recipe ({len(r["dishes_unmatched"])}):'))
            for n in r['dishes_unmatched'][:25]:
                w(f'      {_ascii(n)}')
            if len(r['dishes_unmatched']) > 25:
                w(f'      ...+{len(r["dishes_unmatched"]) - 25} more')
        w(self.style.SUCCESS(
            "\n  next: confirm each option's kind and set pos_mods_string from a real "
            'Lavu "Sales by Item" export before publishing.'))
